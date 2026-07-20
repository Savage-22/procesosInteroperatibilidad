import io
import re
import unicodedata
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlmodel import Session, select

from modules.fichas.application.cambio_service import ETAPAS
from modules.fichas.application.comparacion_service import ComparacionService
from modules.fichas.application.lectura_bd import LecturaBD
from modules.fichas.application.oportunidad_service import (
    clasificar_factibilidad,
    nivel_riesgo,
)
from modules.fichas.application.organizacion_service import OrganizacionService
from modules.fichas.infrastructure.models import (
    AccionCambio,
    Causa,
    FichaIndicador,
    FichaProceso,
    Investigacion,
    Oportunidad,
    Proceso,
    Proyeccion,
)
from modules.procesos.application.proceso_service import ProcesoService
from shared.meses import orden_mes
from shared.semaforo import calcular_semaforo

_AZUL = "1E3654"
_AMARILLO = "F4D100"
_GRIS = "F2F4F7"

# Columnas de la hoja de datos: son exactamente las de la plantilla para que el
# Excel exportado se pueda volver a importar. Las dos últimas son calculadas y
# el importador las ignora (no coinciden con ningún campo que mapee).
_COLUMNAS_DATOS = [
    "Codigo", "Proceso", "Indicador", "Objetivo Estrategico", "Accion Estrategica",
    "Relevancia", "Sentido", "Unidad", "Meta Final", "Año", "Mes",
    "Numerador", "Denominador", "Resultado Esperado", "Resultado Obtenido",
    "Avance T1", "Semaforo",
]


class ExportService:
    """
    Exporta a Excel todo lo que el sistema acumuló en la base de datos.

    El Excel inicial solo trae mediciones; después del trabajo en el sistema hay
    además fichas SIPOC, indicadores caracterizados y el ciclo de mejora completo
    (Ishikawa, oportunidades, Antes/Después y Lewin). Este libro los reúne, con
    la hoja de datos en el mismo formato de la plantilla para poder reimportarlo.
    """

    @staticmethod
    def nombre_archivo(session: Session) -> str:
        org = OrganizacionService.actual(session)
        # La cabecera Content-Disposition debe ser ASCII: se quitan tildes y
        # cualquier carácter que no sea seguro en un nombre de archivo.
        plano = unicodedata.normalize("NFKD", org.nombre or "entidad")
        plano = plano.encode("ascii", "ignore").decode()
        base = re.sub(r"[^A-Za-z0-9]+", "_", plano).strip("_") or "entidad"
        return f"SIIP_{base}_{date.today().isoformat()}.xlsx"

    @staticmethod
    def construir(session: Session) -> bytes:
        wb = Workbook()
        wb.remove(wb.active)

        procesos = {
            p.id: p for p in session.exec(
                select(Proceso).where(Proceso.activo == True)  # noqa: E712
            ).all()
        }

        _hoja(wb, "Resumen", *ExportService._resumen(session, procesos))
        _hoja(wb, "Organizacion", *ExportService._organizacion(session))
        _hoja(wb, "Datos", _COLUMNAS_DATOS, ExportService._datos(session))
        _hoja(wb, "Inventario", *ExportService._inventario(procesos))
        _hoja(wb, "Fichas SIPOC", *ExportService._fichas(session, procesos))
        _hoja(wb, "Indicadores", *ExportService._indicadores(session, procesos))
        _hoja(wb, "Investigaciones", *ExportService._investigaciones(session))
        _hoja(wb, "Ishikawa", *ExportService._causas(session, procesos))
        _hoja(wb, "Oportunidades", *ExportService._oportunidades(session, procesos))
        _hoja(wb, "Proyeccion", *ExportService._proyeccion(session, procesos))
        _hoja(wb, "Antes-Despues", *ExportService._comparacion(session, procesos))
        _hoja(wb, "Gestion del cambio", *ExportService._cambio(session, procesos))

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    # ------------------------------------------------------------------ #
    # Hojas                                                              #
    # ------------------------------------------------------------------ #

    @staticmethod
    def _datos(session: Session) -> list[list]:
        filas = []
        for r in LecturaBD.construir_registros(session):
            avance = ProcesoService.calcular_avance_t1(
                r["resultado_obtenido"], r["resultado_esperado"], r["es_descendente"]
            )
            filas.append([
                r["codigo_proceso"], r["proceso"], r["indicador"],
                r["objetivo_estrategico"], r["accion_estrategica"],
                r["relevancia"], r["sentido"], r["unidad"], r["meta_final"],
                r["anio"], r["mes"], r["numerador"], r["denominador"],
                r["resultado_esperado"], r["resultado_obtenido"],
                round(avance, 2) if avance is not None else None,
                calcular_semaforo(avance),
            ])
        filas.sort(key=lambda f: (f[0], orden_mes(f[10])))
        return filas

    @staticmethod
    def _resumen(session: Session, procesos: dict) -> tuple[list[str], list[list]]:
        """Una fila por proceso con su avance T1 promedio y el estado de su mejora."""
        registros = LecturaBD.construir_registros(session)
        por_codigo: dict[str, list[dict]] = {}
        for r in registros:
            por_codigo.setdefault(r["codigo_proceso"], []).append(r)

        con_causas = _codigos_con(session, Causa, procesos)
        con_oportunidades = _codigos_con(session, Oportunidad, procesos)
        con_cambio = _codigos_con(session, AccionCambio, procesos)

        filas = []
        for codigo in sorted(por_codigo):
            regs = por_codigo[codigo]
            avance = ProcesoService.calcular_promedios(regs)["promedio_avance_t1"]
            filas.append([
                codigo, regs[0]["proceso"], regs[0]["modulo"],
                regs[0]["meta_final"], len(regs), avance, calcular_semaforo(avance),
                "Sí" if codigo in con_causas else "No",
                "Sí" if codigo in con_oportunidades else "No",
                "Sí" if codigo in con_cambio else "No",
            ])
        columnas = [
            "Codigo", "Proceso", "Modulo", "Meta Final", "Meses medidos",
            "Avance T1 promedio", "Semaforo", "Tiene causas", "Tiene oportunidades",
            "Tiene plan de cambio",
        ]
        return columnas, filas

    @staticmethod
    def _organizacion(session: Session) -> tuple[list[str], list[list]]:
        """Identidad de la entidad, para que al reimportar el libro no se pierda."""
        org = OrganizacionService.actual(session)
        return ["Nombre", "Sector"], [[org.nombre, org.sector or ""]]

    @staticmethod
    def _proyeccion(session: Session, procesos: dict) -> tuple[list[str], list[list]]:
        """
        Los meses proyectados tras la mejora, uno por fila. La hoja
        Antes-Despues resume el efecto, pero es calculada; esta trae el dato
        crudo, que es el que se puede volver a importar.
        """
        indicadores = {
            ind.id: ind for ind in session.exec(
                select(FichaIndicador).where(FichaIndicador.activo == True)  # noqa: E712
            ).all()
        }
        filas = []
        for p in session.exec(select(Proyeccion)).all():
            indicador = indicadores.get(p.indicador_id)
            proceso = procesos.get(indicador.proceso_id) if indicador else None
            if proceso is None:
                continue
            for mes in p.meses or []:
                filas.append([
                    proceso.codigo, indicador.nombre, mes.get("mes", ""),
                    mes.get("anio"), mes.get("valor"), p.nota or "",
                ])
        filas.sort(key=lambda f: (f[0], f[1], orden_mes(f[2])))
        return ["Codigo", "Indicador", "Mes", "Año", "Valor Proyectado", "Nota"], filas

    @staticmethod
    def _inventario(procesos: dict) -> tuple[list[str], list[list]]:
        filas = [
            [p.codigo, p.nivel, p.codigo_padre or "", p.nombre, p.producto or "", p.base_legal or ""]
            for p in sorted(procesos.values(), key=lambda x: x.codigo)
        ]
        return ["Codigo", "Nivel", "Codigo Padre", "Nombre", "Producto", "Base Legal"], filas

    @staticmethod
    def _fichas(session: Session, procesos: dict) -> tuple[list[str], list[list]]:
        fichas = session.exec(select(FichaProceso)).all()
        filas = []
        for f in sorted(fichas, key=lambda x: procesos[x.proceso_id].codigo if x.proceso_id in procesos else ""):
            proceso = procesos.get(f.proceso_id)
            if proceso is None:
                continue
            filas.append([
                proceso.codigo, proceso.nombre, f.tipo or "", f.dueno or "",
                f.objetivo or "", f.objetivo_estrategico or "",
                _lista(f.proveedores), _lista(f.entradas), _lista(f.salidas), _lista(f.receptores),
                _lista(f.actividades), _lista(f.riesgos), _lista(f.registros),
                f.elaborado_por or "", f.revisado_por or "", f.aprobado_por or "",
            ])
        columnas = [
            "Codigo", "Proceso", "Tipo", "Dueño", "Objetivo", "Objetivo Estrategico",
            "Proveedores", "Entradas", "Salidas", "Receptores", "Actividades (PDCA)",
            "Riesgos", "Registros", "Elaborado por", "Revisado por", "Aprobado por",
        ]
        return columnas, filas

    @staticmethod
    def _indicadores(session: Session, procesos: dict) -> tuple[list[str], list[list]]:
        indicadores = session.exec(
            select(FichaIndicador).where(FichaIndicador.activo == True)  # noqa: E712
        ).all()
        filas = []
        for ind in indicadores:
            proceso = procesos.get(ind.proceso_id)
            if proceso is None:
                continue
            filas.append([
                proceso.codigo, ind.nombre, ind.tipo or "", ind.sentido, ind.unidad or "",
                ind.formula or "", ind.fuente or "", ind.responsable or "",
                ind.linea_base, ind.meta_final, ind.relevancia,
                ind.objetivo_estrategico or "", ind.accion_estrategica or "",
            ])
        filas.sort(key=lambda f: f[0])
        columnas = [
            "Codigo", "Indicador", "Tipo", "Sentido", "Unidad", "Formula", "Fuente",
            "Responsable", "Linea Base", "Meta Final", "Relevancia",
            "Objetivo Estrategico", "Accion Estrategica",
        ]
        return columnas, filas

    @staticmethod
    def _investigaciones(session: Session) -> tuple[list[str], list[list]]:
        """Sustento académico de cada macroproceso. Cuelga del módulo, no del proceso."""
        investigaciones = session.exec(
            select(Investigacion).where(Investigacion.activo == True)  # noqa: E712
        ).all()
        filas = [
            [
                i.macroproceso, i.titulo, i.autores or "", i.anio, i.tipo or "",
                i.institucion or "", i.url or "", i.aporte or "",
            ]
            for i in investigaciones
        ]
        filas.sort(key=lambda f: (f[0], -(f[3] or 0), f[1]))
        columnas = [
            "Macroproceso", "Titulo", "Autores", "Año", "Tipo", "Institucion",
            "URL", "Aporte al macroproceso",
        ]
        return columnas, filas

    @staticmethod
    def _causas(session: Session, procesos: dict) -> tuple[list[str], list[list]]:
        causas = session.exec(
            select(Causa).where(Causa.activo == True)  # noqa: E712
        ).all()
        filas = []
        for c in causas:
            proceso = procesos.get(c.proceso_id)
            if proceso is None:
                continue
            filas.append([
                proceso.codigo, proceso.nombre, c.categoria, c.descripcion,
                "Sí" if c.es_raiz else "No", c.peso,
            ])
        filas.sort(key=lambda f: (f[0], f[2]))
        return ["Codigo", "Proceso", "Categoria (6M)", "Causa", "Es raiz", "Peso"], filas

    @staticmethod
    def _oportunidades(session: Session, procesos: dict) -> tuple[list[str], list[list]]:
        oportunidades = session.exec(
            select(Oportunidad).where(Oportunidad.activo == True)  # noqa: E712
        ).all()
        filas = []
        for o in oportunidades:
            proceso = procesos.get(o.proceso_id)
            if proceso is None:
                continue
            factibilidad = o.costo * o.impacto
            riesgo = o.probabilidad * o.consecuencia
            filas.append([
                proceso.codigo, o.descripcion, o.accion_propuesta or "",
                o.costo, o.impacto, factibilidad, clasificar_factibilidad(factibilidad),
                o.probabilidad, o.consecuencia, riesgo, nivel_riesgo(riesgo),
                o.estrategia or "", o.estado,
            ])
        filas.sort(key=lambda f: (f[0], -f[5]))
        columnas = [
            "Codigo", "Oportunidad", "Accion Propuesta", "Costo (C)", "Impacto (I)",
            "Factibilidad (F=CxI)", "Plazo", "Probabilidad", "Consecuencia",
            "Riesgo (PxC)", "Nivel de riesgo", "Estrategia", "Estado",
        ]
        return columnas, filas

    @staticmethod
    def _comparacion(session: Session, procesos: dict) -> tuple[list[str], list[list]]:
        filas = []
        for proceso in sorted(procesos.values(), key=lambda p: p.codigo):
            datos = ComparacionService.comparacion(session, proceso.codigo)
            for ind in datos["indicadores"]:
                mejora = ind.get("mejora")
                if not mejora:
                    continue
                filas.append([
                    proceso.codigo, ind["nombre"], ind["meta_final"],
                    mejora["valor_antes"], mejora["semaforo_antes"],
                    mejora["valor_despues"], mejora["semaforo_despues"],
                    mejora["brecha_actual"], mejora["brecha_proyectada"],
                    mejora["mejora_pp"], mejora["mes_alcanza_meta"] or "",
                    ind.get("nota") or "",
                ])
        columnas = [
            "Codigo", "Indicador", "Meta Final", "Valor Antes", "Semaforo Antes",
            "Valor Despues", "Semaforo Despues", "Brecha Actual", "Brecha Proyectada",
            "Mejora (pp)", "Mes que alcanza la meta", "Nota",
        ]
        return columnas, filas

    @staticmethod
    def _cambio(session: Session, procesos: dict) -> tuple[list[str], list[list]]:
        acciones = session.exec(
            select(AccionCambio).where(AccionCambio.activo == True)  # noqa: E712
        ).all()
        filas = []
        for a in acciones:
            proceso = procesos.get(a.proceso_id)
            if proceso is None:
                continue
            filas.append([
                proceso.codigo, a.etapa.capitalize(), a.descripcion,
                a.responsable or "", a.fecha or "", a.estado,
            ])
        orden = {e: i for i, e in enumerate(ETAPAS)}
        filas.sort(key=lambda f: (f[0], orden.get(f[1].lower(), 99)))
        columnas = ["Codigo", "Etapa (Lewin)", "Accion", "Responsable", "Fecha", "Estado"]
        return columnas, filas


# --------------------------------------------------------------------------- #
# Helpers de escritura                                                        #
# --------------------------------------------------------------------------- #

def _lista(valores: list[str] | None) -> str:
    """Une una lista JSON en una celda legible."""
    return " | ".join(valores or [])


def _codigos_con(session: Session, modelo, procesos: dict) -> set[str]:
    """Códigos de proceso que tienen al menos un registro activo del modelo dado."""
    ids = session.exec(
        select(modelo.proceso_id).where(modelo.activo == True)  # noqa: E712
    ).all()
    return {procesos[pid].codigo for pid in ids if pid in procesos}


def _hoja(wb: Workbook, titulo: str, columnas: list[str], filas: list[list]) -> None:
    ws = wb.create_sheet(titulo)
    ws.append(columnas)

    for celda in ws[1]:
        celda.font = Font(bold=True, color=_AMARILLO)
        celda.fill = PatternFill("solid", fgColor=_AZUL)
        celda.alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    for fila in filas:
        ws.append(fila)

    # Ancho por contenido, acotado para que las celdas largas no desborden la hoja
    for i, columna in enumerate(columnas, start=1):
        largo = max(
            [len(str(columna))] + [len(str(f[i - 1])) for f in filas if f[i - 1] is not None]
        )
        ws.column_dimensions[get_column_letter(i)].width = min(max(largo + 2, 10), 55)

    if not filas:
        ws.append(["(sin registros todavía)"])
        ws.cell(row=2, column=1).fill = PatternFill("solid", fgColor=_GRIS)
