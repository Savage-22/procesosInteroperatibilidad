import io

from fastapi import APIRouter
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

router = APIRouter(prefix="/api")

_AZUL = "1E3654"
_AMARILLO = "F4D100"
_GRIS_CLARO = "F2F4F7"
_VERDE_CLARO = "D1FADF"
_ROJO_CLARO = "FFE8E8"

_COLUMNAS = [
    ("Codigo",                  "Código del proceso. Patrón: M<módulo>.<número>  Ej: M1.1, M3.2"),
    ("Proceso",                 "Nombre descriptivo del proceso o servicio que se mide."),
    ("Indicador",               "Descripción del indicador que mide el proceso."),
    ("Objetivo Estrategico",    "(Opcional) Objetivo estratégico al que pertenece este proceso. Ej: OE1 — Mejorar atención al ciudadano"),
    ("Accion Estrategica",      "(Opcional) Acción estratégica dentro del objetivo. Ej: A1.1 — Reducir tiempo de espera"),
    ("Relevancia",         "Importancia del indicador dentro del módulo: 1=Muy relevante  2=Relevante  3=Menos relevante"),
    ("Sentido",            "Dirección del indicador: Ascendente (mayor es mejor)  o  Descendente (menor es mejor, ej. tiempo de espera)."),
    ("Unidad",             "Unidad de medida del resultado. Ejemplos: %  días  horas  atenciones. Usar % o dejar vacío para porcentajes."),
    ("Meta Final",         "Valor meta a alcanzar al cierre del año (diciembre)."),
    ("Año",                "Año al que corresponde el registro. Ej: 2025"),
    ("Mes",                "Mes del registro. Valores válidos: Enero Febrero Marzo Abril Mayo Junio Julio Agosto Septiembre Octubre Noviembre Diciembre"),
    ("Numerador",          "Valor del numerador para calcular el resultado (ej. personas atendidas). Dejar vacío si se usa Resultado Obtenido directamente."),
    ("Denominador",        "Valor del denominador (ej. total de personas). El sistema calcula Resultado = Numerador / Denominador × 100 para unidades porcentuales."),
    ("Resultado Esperado", "Valor que se esperaba obtener en este mes (meta mensual)."),
    ("Resultado Obtenido", "Valor obtenido en el mes. Solo completar si NO se usan Numerador/Denominador. Si ambos están presentes, este campo se ignora."),
]

_EJEMPLOS = [
    # M1 — Diseño de Servicios Públicos Interoperables
    ("M1.1", "Identificación de necesidades del servicio",   "% de requerimientos correctamente identificados",
     "Identificar adecuadamente las necesidades de los usuarios para el diseño de servicios interoperables",
     "Recopilar requerimientos mediante entrevistas y revisión documental",
     1, "Ascendente", "%", 90, 2025, "Enero",   850, 1000, 80, ""),
    ("M1.1", "Identificación de necesidades del servicio",   "% de requerimientos correctamente identificados",
     "Identificar adecuadamente las necesidades de los usuarios para el diseño de servicios interoperables",
     "Recopilar requerimientos mediante entrevistas y revisión documental",
     1, "Ascendente", "%", 90, 2025, "Febrero", 900, 1000, 83, ""),
    ("M1.2", "Análisis del proceso actual",                  "% de procesos analizados completamente (AS-IS)",
     "Analizar integralmente los procesos actuales para identificar oportunidades de mejora",
     "Levantar y documentar información del proceso actual",
     1, "Ascendente", "%", 95, 2025, "Enero",   160, 200,  75, ""),
    ("M1.3", "Diseño del servicio interoperable",            "% de diseños de servicios aprobados sin observaciones",
     "Diseñar servicios interoperables alineados a estándares y necesidades identificadas",
     "Elaborar el diseño del servicio considerando requerimientos y análisis previo",
     1, "Ascendente", "%", 90, 2025, "Enero",   108, 120,  80, ""),
    # M2 — Provisión e Intercambio de Servicios de Información
    ("M2.1", "Gestión de solicitudes de interoperabilidad",  "% de solicitudes correctamente analizadas",
     "Garantizar la calidad de los servicios de interoperabilidad implementados por la entidad",
     "Estandarizar el proceso de análisis de solicitudes de interoperabilidad",
     1, "Ascendente", "%", 95, 2025, "Enero",    95, 100,  85, ""),
    ("M2.2", "Diseño y configuración de integración",        "% de servicios implementados sin errores críticos",
     "Optimizar el tiempo de implementación de los servicios de interoperabilidad",
     "Estandarizar los procesos de diseño, desarrollo y pruebas de interoperabilidad",
     1, "Ascendente", "%", 95, 2025, "Enero",    80, 100,  80, ""),
    ("M2.3", "Ejecución de pruebas de interoperabilidad",    "% de pruebas aprobadas en el primer intento",
     "Asegurar la calidad técnica de los servicios de interoperabilidad durante la fase de pruebas",
     "Fortalecer los procedimientos de pruebas técnicas de interoperabilidad",
     1, "Ascendente", "%", 90, 2025, "Enero",    70, 100,  75, ""),
    # M3 — Integración de Sistemas y Plataformas Interoperables
    ("M3.1", "Análisis y especificación de la integración",  "% de especificaciones técnicas aprobadas",
     "Definir las reglas técnicas y de negocio necesarias para la conexión entre sistemas",
     "Relevar, documentar y validar los requerimientos de datos a integrar",
     1, "Ascendente", "%", 90, 2025, "Enero",   170, 200,  70, ""),
    ("M3.2", "Desarrollo y configuración del middleware/API", "% de componentes de software sin vulnerabilidades",
     "Construir los componentes de software que ejecutan la integración entre plataformas",
     "Codificar, configurar y securizar los servicios de API y middleware",
     1, "Ascendente", "%", 90, 2025, "Enero",   108, 120,  80, ""),
    ("M3.3", "Pruebas y aseguramiento de interoperabilidad", "% de pruebas exitosas de interoperabilidad",
     "Validar que los datos se transmiten correctamente y que el sistema cumple los SLA definidos",
     "Ejecutar pruebas funcionales, de estrés y verificar la estabilidad del servicio integrado",
     1, "Ascendente", "%", 90, 2025, "Enero",    90, 100,  80, ""),
    # M4 — Orquestación de Servicios Interoperables
    ("M4.1", "Planificación de la orquestación",             "Nivel de cumplimiento del plan de orquestación (%)",
     "Planificar la orquestación eficiente de servicios interoperables",
     "Identificar y analizar los servicios interoperables requeridos",
     1, "Ascendente", "%", 100, 2025, "Enero",   60, 100,  75, ""),
    ("M4.2", "Configuración de servicios interoperables",    "% de servicios configurados e integrados correctamente",
     "Configurar servicios interoperables de manera eficiente",
     "Configurar servicios en plataformas de integración",
     1, "Ascendente", "%", 95, 2025, "Enero",    70, 100,  70, ""),
    ("M4.3", "Ejecución y disponibilidad de servicios",      "% de disponibilidad de servicios interoperables",
     "Ejecutar los servicios interoperables de manera eficiente",
     "Ejecutar los servicios interoperables configurados",
     1, "Ascendente", "%", 95, 2025, "Enero",    82, 100,  78, ""),
]

_ENCABEZADOS = [c[0] for c in _COLUMNAS]

# --------------------------------------------------------------------------- #
# Hojas del resto del estado del sistema                                       #
#                                                                              #
# La hoja Datos solo cubre las mediciones. Todo lo que se trabaja dentro del   #
# sistema —inventario, fichas SIPOC, indicadores caracterizados y el ciclo de  #
# mejora— vive en estas hojas, con exactamente los mismos encabezados que      #
# escribe el export: así el libro exportado se puede volver a subir entero y   #
# una entidad nueva puede llegar con todo ya cargado.                          #
# --------------------------------------------------------------------------- #

_HOJAS_ESTADO = [
    (
        "Organizacion",
        "Identidad de la entidad. Una sola fila.",
        [
            ("Nombre", "Nombre oficial de la entidad. Ej: Municipalidad Distrital de Miraflores"),
            ("Sector", "Sector al que pertenece. Ej: Gobierno local, Salud, Educación"),
        ],
        [("Municipalidad Distrital de Miraflores", "Gobierno local")],
    ),
    (
        "Inventario",
        "Anexo 1 — inventario de procesos. Opcional: los procesos que aparezcan en "
        "la hoja Datos se crean solos. Úsala para darles producto, base legal y jerarquía.",
        [
            ("Codigo", "Código del proceso. Ej: M1, M1.1, M1.1.2"),
            ("Nivel", "0 = macroproceso, 1 = proceso, 2 = subproceso, 3 = actividad"),
            ("Codigo Padre", "Código del proceso de nivel superior. Vacío en los macroprocesos"),
            ("Nombre", "Nombre del proceso"),
            ("Producto", "Producto o servicio que entrega el proceso"),
            ("Base Legal", "Norma que sustenta el proceso. Ej: D.S. 123-2024-PCM"),
        ],
        [
            ("M1", 0, "", "Diseño de servicios públicos interoperables", "", ""),
            ("M1.1", 1, "M1", "Identificación de necesidades del servicio",
             "Informe de requerimientos validado", "Directiva N° 0056-2024-CEPLAN/PCD"),
        ],
    ),
    (
        "Fichas SIPOC",
        "Anexo 2 — caracterización de cada proceso. En las columnas de lista, separa "
        "los elementos con una barra vertical: Uno | Dos | Tres",
        [
            ("Codigo", "Código del proceso que se caracteriza. Ej: M1.1"),
            ("Tipo", "misional | estratégico | soporte"),
            ("Dueño", "Responsable de que el proceso cumpla su objetivo (process owner)"),
            ("Objetivo", "Para qué existe el proceso, en una frase"),
            ("Objetivo Estrategico", "Objetivo estratégico institucional al que aporta"),
            ("Proveedores", "Quién entrega las entradas. Separar con |"),
            ("Entradas", "Qué recibe el proceso. Separar con |"),
            ("Salidas", "Qué produce el proceso. Separar con |"),
            ("Receptores", "Quién recibe las salidas. Separar con |"),
            ("Actividades (PDCA)", "Actividades del ciclo Planificar-Hacer-Verificar-Actuar. Separar con |"),
            ("Riesgos", "Riesgos del proceso. Separar con |"),
            ("Registros", "Documentos o sistemas donde queda evidencia. Separar con |"),
            ("Elaborado por", "Quién elaboró la ficha"),
            ("Revisado por", "Quién la revisó"),
            ("Aprobado por", "Quién la aprobó"),
        ],
        [(
            "M1.1", "misional", "Jefe de Planeamiento",
            "Identificar las necesidades de los usuarios del servicio",
            "Mejorar la atención al ciudadano",
            "Área usuaria | Mesa de partes", "Solicitud de servicio | Normativa vigente",
            "Informe de requerimientos", "Área de procesos | Alta dirección",
            "Planificar el levantamiento | Entrevistar a usuarios | Validar el informe | Ajustar el instrumento",
            "Requerimientos incompletos | Rotación del personal entrevistador",
            "Acta de reunión | Sistema de trámite documentario",
            "Analista de procesos", "Jefe de Planeamiento", "Gerencia Municipal",
        )],
    ),
    (
        "Indicadores",
        "Anexo 4 — caracterización de los indicadores. Complementa la hoja Datos: allí "
        "se registran las mediciones y aquí la ficha del indicador. El nombre del "
        "indicador debe coincidir exactamente con el de la hoja Datos.",
        [
            ("Codigo", "Código del proceso al que pertenece. Ej: M1.1"),
            ("Indicador", "Nombre del indicador, idéntico al de la hoja Datos"),
            ("Tipo", "eficacia | eficiencia | efectividad"),
            ("Sentido", "Ascendente (mayor es mejor) | Descendente (menor es mejor)"),
            ("Unidad", "% | días | horas | atenciones…"),
            ("Formula", "Cómo se calcula. Ej: requerimientos validados / requerimientos totales × 100"),
            ("Fuente", "De dónde sale el dato"),
            ("Responsable", "Quién reporta la medición"),
            ("Linea Base", "Valor de partida antes de la intervención"),
            ("Meta Final", "Valor a alcanzar al cierre del año"),
            ("Relevancia", "1 = muy relevante, 2 = relevante, 3 = menos relevante"),
            ("Objetivo Estrategico", "Objetivo estratégico al que aporta"),
            ("Accion Estrategica", "Acción estratégica dentro de ese objetivo"),
        ],
        [(
            "M1.1", "% de requerimientos correctamente identificados", "eficacia",
            "Ascendente", "%", "requerimientos validados / requerimientos totales × 100",
            "Sistema de trámite documentario", "Jefe de Planeamiento", 60, 90, 1,
            "Mejorar la atención al ciudadano", "Estandarizar el levantamiento de requerimientos",
        )],
    ),
    (
        "Ishikawa",
        "Mejora I — causas del bajo desempeño, clasificadas con las 6M. Marca como "
        "causa raíz las que conviene atacar.",
        [
            ("Codigo", "Código del proceso. Ej: M1.1"),
            ("Categoria (6M)", "Método | Personas | Entorno | Medición | Máquina-TI | Materiales"),
            ("Causa", "La causa, concreta"),
            ("Es raiz", "Sí | No — marca Sí si es la causa que conviene atacar"),
            ("Peso", "Frecuencia o impacto de la causa, de 1 a 10. Ordena el Pareto de causas"),
        ],
        [
            ("M1.1", "Método", "No existe un procedimiento estándar de levantamiento", "Sí", 8),
            ("M1.1", "Personas", "El personal entrevistador no está capacitado", "No", 5),
        ],
    ),
    (
        "Oportunidades",
        "Mejora II — oportunidades priorizadas. La factibilidad (F = Costo × Impacto) y "
        "el nivel de riesgo (Probabilidad × Consecuencia) los calcula el sistema.",
        [
            ("Codigo", "Código del proceso. Ej: M1.1"),
            ("Oportunidad", "La oportunidad de mejora"),
            ("Accion Propuesta", "Qué hacer exactamente"),
            ("Costo (C)", "1 a 5, donde 1 es el costo más bajo"),
            ("Impacto (I)", "1 a 5, donde 5 es el impacto más alto"),
            ("Probabilidad", "1 a 5 — probabilidad de que el riesgo ocurra"),
            ("Consecuencia", "1 a 5 — gravedad si ocurre"),
            ("Estrategia", "evitar | mitigar | transferir | aceptar"),
            ("Estado", "propuesta | en_curso | implementada | descartada"),
        ],
        [(
            "M1.1", "Estandarizar el instrumento de levantamiento de requerimientos",
            "Elaborar y aprobar una guía única con formato de entrevista",
            2, 5, 3, 4, "mitigar", "en_curso",
        )],
    ),
    (
        "Investigaciones",
        "Sustento académico del macroproceso: tesis, artículos o normas que respaldan "
        "que el proceso está bien planteado. Cuelga del módulo (M1, M2…), no de un "
        "proceso concreto. Conviene al menos una por macroproceso.",
        [
            ("Macroproceso", "Módulo que sustenta. Patrón M<número>. Ej: M1, M2, M3, M4"),
            ("Titulo", "Título completo de la investigación"),
            ("Autores", "Autor o autores, como aparecen en la fuente"),
            ("Año", "Año de publicación. Ej: 2024"),
            ("Tipo", "tesis | artículo | libro | informe | norma"),
            ("Institucion", "Universidad, entidad o revista que la publica"),
            ("URL", "Enlace al repositorio. Debe empezar con http:// o https://"),
            ("Aporte al macroproceso", "Qué del macroproceso respalda esta investigación"),
        ],
        [
            ("M1", "Modelo de plataforma de interoperabilidad basado en arquitectura orientada "
             "a servicios para la transacción automática de datos entre entidades públicas",
             "Vílchez Inga, César", 2020, "tesis",
             "Universidad Nacional Federico Villarreal",
             "https://repositorio.unfv.edu.pe/items/7f008a15-42a8-40df-a869-ec2ff0ccd5b0",
             "Sustenta el diseño del servicio interoperable con estándares SOA (M1.3)"),
            ("M2", "Mejoras en la implementación del proceso de interoperabilidad en las "
             "entidades públicas",
             "Salas Palomino, Fiorella; Giles Magallanes, Rosa", 2025, "tesis",
             "Pontificia Universidad Católica del Perú",
             "https://tesis.pucp.edu.pe/items/6bbc5f2a-feb0-4d90-86c7-5e507d423e91",
             "Identifica las causas del fracaso en la implementación vía PIDE (M2.1 y M2.2)"),
            ("M3", "Implementación de un sistema de interoperabilidad de información clínica "
             "basado en los estándares internacionales HL7 FHIR y openEHR",
             "Porras Gamarra, Hernán Johel", 2024, "tesis",
             "Universidad Nacional Federico Villarreal",
             "https://repositorio.unfv.edu.pe/handle/20.500.13084/9034",
             "Caso de especificación técnica, APIs y pruebas de integración (M3.1 a M3.3)"),
            ("M4", "Orquestación de servicios web utilizando Oracle BPEL Process Manager. "
             "Aplicación a un caso práctico",
             "Durango, José Ignacio", None, "tesis", "Escuela Politécnica Nacional",
             "https://bibdigital.epn.edu.ec/handle/15000/542",
             "Respalda la planificación y configuración de la orquestación (M4.1 y M4.2)"),
        ],
    ),
    (
        "Proyeccion",
        "Mejora III — valor que se espera alcanzar mes a mes tras aplicar la mejora. "
        "Una fila por mes proyectado. Alimenta la comparación Antes/Después.",
        [
            ("Codigo", "Código del proceso. Ej: M1.1"),
            ("Indicador", "Nombre del indicador proyectado, idéntico al de la hoja Datos"),
            ("Mes", "Mes proyectado. Enero … Diciembre"),
            ("Año", "Año del mes proyectado. Ej: 2026"),
            ("Valor Proyectado", "Valor que se espera obtener ese mes tras la mejora"),
            ("Nota", "Por qué esa proyección es alcanzable"),
        ],
        [
            ("M1.1", "% de requerimientos correctamente identificados", "Junio", 2026, 85,
             "La guía estandarizada entra en vigencia en junio"),
            ("M1.1", "% de requerimientos correctamente identificados", "Julio", 2026, 90, ""),
        ],
    ),
    (
        "Gestion del cambio",
        "Mejora IV — plan de gestión del cambio según el modelo de Kurt Lewin. "
        "Conviene al menos una acción por etapa.",
        [
            ("Codigo", "Código del proceso. Ej: M1.1"),
            ("Etapa (Lewin)", "descongelar | cambiar | recongelar"),
            ("Accion", "Acción concreta del plan"),
            ("Responsable", "Rol o persona responsable"),
            ("Fecha", "Fecha compromiso en formato AAAA-MM-DD. Ej: 2026-08-31"),
            ("Estado", "pendiente | en_curso | hecho"),
        ],
        [
            ("M1.1", "descongelar", "Presentar el diagnóstico de causas al equipo",
             "Jefe de Planeamiento", "2026-06-30", "hecho"),
            ("M1.1", "cambiar", "Capacitar al equipo en la nueva guía",
             "Analista de procesos", "2026-08-31", "en_curso"),
            ("M1.1", "recongelar", "Incorporar la guía al manual de procedimientos",
             "Gerencia Municipal", "2026-10-31", "pendiente"),
        ],
    ),
]


def _borde_fino():
    lado = Side(style="thin", color="CCCCCC")
    return Border(left=lado, right=lado, top=lado, bottom=lado)


def _hoja_datos(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "Datos"

    fill_header  = PatternFill("solid", fgColor=_AZUL)
    fill_ejemplo = PatternFill("solid", fgColor=_GRIS_CLARO)
    font_header  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    font_ejemplo = Font(name="Calibri", italic=True, color="555555", size=10)
    font_normal  = Font(name="Calibri", size=10)
    alin_centro  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    alin_izq     = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    # — Fila 1: encabezados —
    for col_idx, nombre in enumerate(_ENCABEZADOS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=nombre)
        cell.fill   = fill_header
        cell.font   = font_header
        cell.border = _borde_fino()
        cell.alignment = alin_centro

    # — Filas de ejemplo —
    for row_idx, fila in enumerate(_EJEMPLOS, start=2):
        for col_idx, valor in enumerate(fila, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=valor if valor != "" else None)
            cell.fill      = fill_ejemplo
            cell.font      = font_ejemplo
            cell.border    = _borde_fino()
            cell.alignment = alin_izq

    # — Fila vacía para que el usuario empiece a llenar —
    fila_nueva = len(_EJEMPLOS) + 2
    for col_idx in range(1, len(_ENCABEZADOS) + 1):
        cell = ws.cell(row=fila_nueva, column=col_idx)
        cell.border    = _borde_fino()
        cell.font      = font_normal
        cell.alignment = alin_izq

    # — Anchos de columna —
    anchos = [10, 32, 38, 30, 28, 11, 13, 8, 12, 6, 13, 13, 13, 18, 18]
    for i, ancho in enumerate(anchos, start=1):
        ws.column_dimensions[get_column_letter(i)].width = ancho

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"


def _hoja_estado(wb: Workbook, nombre: str, ayuda: str, columnas: list, ejemplos: list) -> None:
    """
    Una hoja del estado del sistema: fila de ayuda, encabezados y ejemplos en
    gris. El mismo formato que la hoja Datos, para que se lean todas igual.
    """
    ws = wb.create_sheet(nombre)

    font_ayuda = Font(name="Calibri", italic=True, color="555555", size=10)
    alin_izq   = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columnas))
    celda_ayuda = ws.cell(row=1, column=1, value=ayuda)
    celda_ayuda.fill      = PatternFill("solid", fgColor="FFFDE7")
    celda_ayuda.font      = font_ayuda
    celda_ayuda.alignment = alin_izq
    ws.row_dimensions[1].height = 34

    for col_idx, (encabezado, _) in enumerate(columnas, start=1):
        cell = ws.cell(row=2, column=col_idx, value=encabezado)
        cell.fill      = PatternFill("solid", fgColor=_AZUL)
        cell.font      = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        cell.border    = _borde_fino()
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 28

    for row_idx, fila in enumerate(ejemplos, start=3):
        for col_idx, valor in enumerate(fila, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=valor if valor != "" else None)
            cell.fill      = PatternFill("solid", fgColor=_GRIS_CLARO)
            cell.font      = Font(name="Calibri", italic=True, color="555555", size=10)
            cell.border    = _borde_fino()
            cell.alignment = alin_izq

    # Fila vacía lista para escribir
    for col_idx in range(1, len(columnas) + 1):
        cell = ws.cell(row=len(ejemplos) + 3, column=col_idx)
        cell.border    = _borde_fino()
        cell.font      = Font(name="Calibri", size=10)
        cell.alignment = alin_izq

    for i, (encabezado, _) in enumerate(columnas, start=1):
        largo = max(
            [len(encabezado)]
            + [len(str(f[i - 1])) for f in ejemplos if f[i - 1] not in (None, "")]
        )
        ws.column_dimensions[get_column_letter(i)].width = min(max(largo + 2, 12), 45)

    ws.freeze_panes = "A3"


def _hoja_instrucciones(wb: Workbook) -> None:
    ws = wb.create_sheet("Instrucciones")

    fill_titulo  = PatternFill("solid", fgColor=_AZUL)
    fill_col     = PatternFill("solid", fgColor="DCE8F5")
    fill_nota    = PatternFill("solid", fgColor="FFFDE7")
    fill_verde   = PatternFill("solid", fgColor=_VERDE_CLARO)
    fill_rojo    = PatternFill("solid", fgColor=_ROJO_CLARO)
    font_titulo  = Font(name="Calibri", bold=True, color="FFFFFF", size=13)
    font_sub     = Font(name="Calibri", bold=True, color=_AZUL, size=11)
    font_campo   = Font(name="Calibri", bold=True, size=10)
    font_normal  = Font(name="Calibri", size=10)
    font_nota    = Font(name="Calibri", italic=True, color="555555", size=10)
    alin_izq     = Alignment(horizontal="left", vertical="top", wrap_text=True)
    alin_centro  = Alignment(horizontal="center", vertical="center")

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 70

    # — Título —
    ws.merge_cells("A1:B1")
    cell = ws["A1"]
    cell.value     = "Instrucciones para completar la plantilla — Sistema Dashboard CEPLAN"
    cell.fill      = fill_titulo
    cell.font      = font_titulo
    cell.alignment = alin_centro
    ws.row_dimensions[1].height = 30

    # — Subtítulo —
    ws.merge_cells("A2:B2")
    cell = ws["A2"]
    cell.value     = "Directiva N° 0056-2024-CEPLAN/PCD — Seguimiento y evaluación de procesos"
    cell.font      = Font(name="Calibri", italic=True, color="666666", size=10)
    cell.alignment = alin_centro
    ws.row_dimensions[2].height = 18

    fila = 4

    def encabezado_tabla(izquierda: str, derecha: str) -> None:
        nonlocal fila
        for columna, texto in ((1, izquierda), (2, derecha)):
            cell = ws.cell(row=fila, column=columna, value=texto)
            cell.fill      = PatternFill("solid", fgColor="2D4F7A")
            cell.font      = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
            cell.alignment = alin_centro
            cell.border    = _borde_fino()
        ws.row_dimensions[fila].height = 20
        fila += 1

    def filas_columnas(columnas, alto: int = 42) -> None:
        nonlocal fila
        for nombre, descripcion in columnas:
            cell_a = ws.cell(row=fila, column=1, value=nombre)
            cell_b = ws.cell(row=fila, column=2, value=descripcion)
            cell_a.fill      = fill_col
            cell_a.font      = font_campo
            cell_b.font      = font_normal
            cell_a.alignment = alin_izq
            cell_b.alignment = alin_izq
            cell_a.border    = _borde_fino()
            cell_b.border    = _borde_fino()
            ws.row_dimensions[fila].height = alto
            fila += 1

    def subtitulo(texto: str, ayuda: str | None = None) -> None:
        nonlocal fila
        ws.merge_cells(f"A{fila}:B{fila}")
        cell = ws[f"A{fila}"]
        cell.value     = texto
        cell.font      = font_sub
        cell.alignment = alin_izq
        ws.row_dimensions[fila].height = 20
        fila += 1
        if ayuda:
            ws.merge_cells(f"A{fila}:B{fila}")
            cell = ws[f"A{fila}"]
            cell.value     = ayuda
            cell.font      = font_nota
            cell.alignment = alin_izq
            ws.row_dimensions[fila].height = 32
            fila += 1

    # — Hoja Datos: las mediciones mensuales —
    subtitulo(
        "Hoja «Datos» — mediciones mensuales",
        "Es la única hoja obligatoria. Con ella el dashboard, el tablero, el Pareto y "
        "las predicciones ya funcionan.",
    )
    encabezado_tabla("Columna", "Descripción y valores válidos")
    filas_columnas(_COLUMNAS)

    fila += 1

    # — Resto de hojas: el estado que se trabaja dentro del sistema —
    subtitulo(
        "Hojas opcionales — el resto del estado del sistema",
        "Todas son opcionales y se pueden llenar por partes. Son las mismas hojas que "
        "exporta el botón «Exportar Excel», con los mismos encabezados: el libro "
        "exportado se puede volver a subir tal cual y nada se pierde.",
    )
    for nombre_hoja, ayuda_hoja, columnas_hoja, _ in _HOJAS_ESTADO:
        subtitulo(f"Hoja «{nombre_hoja}»", ayuda_hoja)
        encabezado_tabla("Columna", "Descripción y valores válidos")
        filas_columnas(columnas_hoja, alto=34)
        fila += 1

    # — Notas importantes —
    subtitulo("Notas importantes")

    notas = [
        ("Actualizar sin perder lo trabajado",
         "Si ya usas el sistema, descarga tu libro con «Exportar Excel», edítalo y vuelve a subirlo: "
         "trae todas estas hojas ya llenas con tus datos. La importación es idempotente —se actualiza "
         "por código de proceso, nombre de indicador y mes— así que resubir el mismo archivo no duplica nada."),
        ("Columnas calculadas",
         "El libro exportado trae además columnas que el sistema calcula (Avance T1, Semaforo, "
         "Factibilidad, Nivel de riesgo, la hoja Resumen y la hoja Antes-Despues). Al reimportar se "
         "ignoran: no hace falta borrarlas ni mantenerlas al día."),
        ("Orden de llenado",
         "Empieza por la hoja Datos: los procesos e indicadores que aparezcan allí se crean solos y "
         "las demás hojas se cuelgan de ellos por su Codigo. Una fila cuyo Codigo no exista se omite."),
        ("Numerador / Denominador vs Resultado Obtenido",
         "Si completas Numerador y Denominador, el sistema calcula automáticamente el Resultado Obtenido "
         "(Numerador/Denominador × 100 para unidades %). El campo Resultado Obtenido solo se usa si no hay "
         "numerador/denominador disponibles (ej. indicadores de días promedio)."),
        ("Código del proceso",
         "Debe seguir el patrón M<n>.<m> donde <n> es el número de módulo y <m> el número de proceso dentro "
         "del módulo. Ejemplos válidos: M1.1  M2.3  M3.2  M4.1  M5.1"),
        ("Módulos",
         "El sistema espera los módulos M1, M2, M3 y M4 por defecto. Puedes agregar M5, M6, etc. configurando "
         "la variable de entorno MODULOS_ESPERADOS en el servidor."),
        ("Filas de ejemplo",
         "Las filas en gris de la hoja Datos son ejemplos. Puedes reemplazarlas o eliminarlas. "
         "El sistema solo procesa filas cuyo Codigo coincida con el patrón M#.# y cuyo Mes sea un mes válido."),
    ]

    for titulo_nota, texto_nota in notas:
        cell_a = ws.cell(row=fila, column=1, value=titulo_nota)
        cell_b = ws.cell(row=fila, column=2, value=texto_nota)
        cell_a.fill      = fill_nota
        cell_b.fill      = fill_nota
        cell_a.font      = Font(name="Calibri", bold=True, size=10, color="7A5C00")
        cell_b.font      = Font(name="Calibri", size=10, color="555555")
        cell_a.alignment = alin_izq
        cell_b.alignment = alin_izq
        cell_a.border    = _borde_fino()
        cell_b.border    = _borde_fino()
        ws.row_dimensions[fila].height = 60
        fila += 1


@router.get("/plantilla", response_class=StreamingResponse)
def descargar_plantilla():
    wb = Workbook()
    _hoja_datos(wb)
    _hoja_instrucciones(wb)
    for nombre, ayuda, columnas, ejemplos in _HOJAS_ESTADO:
        _hoja_estado(wb, nombre, ayuda, columnas, ejemplos)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=plantilla_ceplan.xlsx"},
    )
