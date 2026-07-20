import io

import pytest
from openpyxl import load_workbook
from sqlmodel import Session, SQLModel, create_engine

from modules.export.application.export_service import ExportService
from modules.fichas.application.errores import ErrorNoEncontrado, ErrorValidacion
from modules.fichas.application.importador_estado import ImportadorEstado
from modules.fichas.application.inventario_service import InventarioService
from modules.fichas.application.investigacion_service import InvestigacionService
from modules.fichas.application.organizacion_service import OrganizacionService

UNFV = {
    "macroproceso": "M1",
    "titulo": "Modelo de plataforma de interoperabilidad basado en SOA",
    "autores": "Vílchez Inga, César",
    "anio": 2020,
    "tipo": "tesis",
    "institucion": "Universidad Nacional Federico Villarreal",
    "url": "https://repositorio.unfv.edu.pe/items/7f008a15",
    "aporte": "Sustenta el diseño del servicio interoperable",
}


@pytest.fixture
def session():
    engine = create_engine("sqlite://", connect_args={"check_same_thread": False})
    SQLModel.metadata.create_all(engine)
    with Session(engine) as s:
        yield s


# ── Registro y agrupación ─────────────────────────────────────────────────────

def test_registra_una_investigacion_del_macroproceso(session):
    creada = InvestigacionService.crear(session, UNFV)
    assert creada["macroproceso"] == "M1"
    assert creada["anio"] == 2020

    datos = InvestigacionService.listar(session)
    assert [i["titulo"] for i in datos["investigaciones"]["M1"]] == [UNFV["titulo"]]
    assert datos["total"] == 1


def test_los_cuatro_macroprocesos_aparecen_aunque_no_tengan_sustento(session):
    datos = InvestigacionService.listar(session)
    assert datos["macroprocesos"] == ["M1", "M2", "M3", "M4"]
    assert datos["sin_sustento"] == ["M1", "M2", "M3", "M4"]


def test_el_macroproceso_del_inventario_se_suma_a_los_esperados(session):
    InventarioService.crear(session, {"codigo": "M5.1", "nombre": "Nuevo módulo"})
    assert "M5" in InvestigacionService.listar(session)["macroprocesos"]


def test_acepta_el_codigo_de_un_proceso_y_lo_reduce_a_su_macroproceso(session):
    creada = InvestigacionService.crear(session, {**UNFV, "macroproceso": "m2.3"})
    assert creada["macroproceso"] == "M2"


def test_ordena_las_investigaciones_de_la_mas_reciente_a_la_mas_antigua(session):
    InvestigacionService.crear(session, {**UNFV, "titulo": "Antigua", "anio": 2015})
    InvestigacionService.crear(session, {**UNFV, "titulo": "Reciente", "anio": 2025})
    titulos = [i["titulo"] for i in InvestigacionService.listar(session)["investigaciones"]["M1"]]
    assert titulos == ["Reciente", "Antigua"]


def test_filtra_por_macroproceso(session):
    InvestigacionService.crear(session, UNFV)
    InvestigacionService.crear(session, {**UNFV, "macroproceso": "M2", "titulo": "Otra"})
    datos = InvestigacionService.listar(session, "M2")
    assert datos["total"] == 1
    assert datos["investigaciones"]["M2"][0]["titulo"] == "Otra"
    assert datos["investigaciones"]["M1"] == []


# ── Validación ────────────────────────────────────────────────────────────────

@pytest.mark.parametrize("campo,valor", [
    ("macroproceso", "modulo 1"),
    ("titulo", "   "),
    ("tipo", "podcast"),
    ("url", "repositorio.unfv.edu.pe"),
    ("anio", 12),
])
def test_rechaza_los_campos_invalidos(session, campo, valor):
    with pytest.raises(ErrorValidacion):
        InvestigacionService.crear(session, {**UNFV, campo: valor})


def test_actualizar_y_eliminar(session):
    creada = InvestigacionService.crear(session, UNFV)
    actualizada = InvestigacionService.actualizar(session, creada["id"], {"anio": 2021})
    assert actualizada["anio"] == 2021

    InvestigacionService.eliminar(session, creada["id"])
    assert InvestigacionService.listar(session)["total"] == 0
    # El borrado es lógico, así que volver a tocarla es un 404, no un registro fantasma
    with pytest.raises(ErrorNoEncontrado):
        InvestigacionService.actualizar(session, creada["id"], {"anio": 2022})


# ── Round-trip por Excel ──────────────────────────────────────────────────────

def test_la_hoja_de_investigaciones_va_y_vuelve_en_el_excel(session):
    InvestigacionService.crear(session, UNFV)
    contenido = ExportService.construir(session)

    hoja = load_workbook(io.BytesIO(contenido))["Investigaciones"]
    fila = list(hoja.iter_rows(min_row=2, values_only=True))[0]
    assert fila[0] == "M1"
    assert fila[3] == 2020

    engine = create_engine("sqlite://", connect_args={"check_same_thread": False})
    SQLModel.metadata.create_all(engine)
    with Session(engine) as destino:
        ImportadorEstado.importar(
            destino, io.BytesIO(contenido), OrganizacionService.actual(destino)
        )
        importada = InvestigacionService.listar(destino)["investigaciones"]["M1"][0]
        assert importada["titulo"] == UNFV["titulo"]
        assert importada["url"] == UNFV["url"]
        assert importada["tipo"] == "tesis"
        assert importada["aporte"] == UNFV["aporte"]


def test_reimportar_el_mismo_libro_no_duplica(session):
    InvestigacionService.crear(session, UNFV)
    contenido = ExportService.construir(session)
    org = OrganizacionService.actual(session)

    ImportadorEstado.importar(session, io.BytesIO(contenido), org)
    ImportadorEstado.importar(session, io.BytesIO(contenido), org)
    assert InvestigacionService.listar(session)["total"] == 1
