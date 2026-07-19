import io

import pytest
from openpyxl import load_workbook
from sqlmodel import Session, SQLModel, create_engine

from modules.export.application.export_service import ExportService
from modules.fichas.application.cambio_service import CambioService
from modules.fichas.application.causa_service import CausaService
from modules.fichas.application.comparacion_service import ComparacionService
from modules.fichas.application.indicador_service import IndicadorService
from modules.fichas.application.inventario_service import InventarioService
from modules.fichas.application.oportunidad_service import OportunidadService
from modules.fichas.application.organizacion_service import OrganizacionService
from modules.procesos.infrastructure.excel_reader import parsear_excel


@pytest.fixture
def session():
    engine = create_engine("sqlite://", connect_args={"check_same_thread": False})
    SQLModel.metadata.create_all(engine)
    with Session(engine) as s:
        yield s


@pytest.fixture
def proceso_completo(session):
    """Un proceso con medición y las cuatro partes de la mejora registradas."""
    InventarioService.crear(session, {"codigo": "M1.1", "nombre": "Identificación"})
    ind = IndicadorService.crear(session, "M1.1", {
        "nombre": "% de requerimientos", "unidad": "%", "meta_final": 90.0,
    })
    IndicadorService.guardar_medicion(session, ind["id"], {
        "mes": "Enero", "anio": 2025, "numerador": 8, "denominador": 10,
        "resultado_esperado": 75,
    })
    CausaService.crear(session, "M1.1", {
        "categoria": "Método", "descripcion": "Sin procedimiento", "es_raiz": True, "peso": 8,
    })
    OportunidadService.crear(session, "M1.1", {
        "descripcion": "Estandarizar", "costo": 2, "impacto": 5,
        "probabilidad": 3, "consecuencia": 4, "estrategia": "mitigar",
    })
    CambioService.crear(session, "M1.1", {
        "etapa": "descongelar", "descripcion": "Sensibilizar al equipo",
    })
    ComparacionService.guardar_proyeccion(session, ind["id"], {
        "meses": [{"mes": "Diciembre", "anio": 2025, "valor": 90}],
        "nota": "Con el procedimiento estandarizado",
    })
    return session


def _abrir(session) -> dict:
    wb = load_workbook(io.BytesIO(ExportService.construir(session)))
    return {nombre: wb[nombre] for nombre in wb.sheetnames}


def _filas(ws) -> list[tuple]:
    return list(ws.iter_rows(min_row=2, values_only=True))


# ── Estructura ────────────────────────────────────────────────────────────────

def test_export_incluye_una_hoja_por_parte_del_trabajo(session, proceso_completo):
    hojas = _abrir(session)
    assert {
        "Resumen", "Datos", "Inventario", "Fichas SIPOC", "Indicadores",
        "Ishikawa", "Oportunidades", "Antes-Despues", "Gestion del cambio",
    } <= set(hojas)


def test_export_sin_datos_no_falla(session):
    hojas = _abrir(session)
    # Cada hoja se emite igual, marcada como vacía, para que el libro sea uniforme
    assert _filas(hojas["Ishikawa"]) == [("(sin registros todavía)", None, None, None, None, None)]


# ── Contenido de la mejora ────────────────────────────────────────────────────

def test_export_lleva_las_causas_ishikawa(session, proceso_completo):
    fila = _filas(_abrir(session)["Ishikawa"])[0]
    assert fila[0] == "M1.1"
    assert fila[2] == "Método"
    assert fila[4] == "Sí"  # es raíz


def test_export_calcula_factibilidad_y_riesgo_de_la_oportunidad(session, proceso_completo):
    fila = _filas(_abrir(session)["Oportunidades"])[0]
    assert fila[3] == 2 and fila[4] == 5      # costo, impacto
    assert fila[5] == 10                       # F = C × I
    assert fila[7] == 3 and fila[8] == 4      # probabilidad, consecuencia
    assert fila[9] == 12                       # riesgo = P × C


def test_export_reporta_el_antes_y_despues(session, proceso_completo):
    fila = _filas(_abrir(session)["Antes-Despues"])[0]
    assert fila[0] == "M1.1"
    assert fila[3] == 80.0   # valor antes (8/10)
    assert fila[5] == 90.0   # valor después (proyectado)


def test_export_lleva_el_plan_de_cambio_lewin(session, proceso_completo):
    fila = _filas(_abrir(session)["Gestion del cambio"])[0]
    assert fila[1] == "Descongelar"
    assert fila[2] == "Sensibilizar al equipo"


# ── Round-trip ────────────────────────────────────────────────────────────────

def test_la_hoja_de_datos_se_puede_reimportar(session, proceso_completo):
    contenido = ExportService.construir(session)
    registros, advertencias = parsear_excel(io.BytesIO(contenido))

    assert len(registros) == 1
    registro = registros[0]
    assert registro["codigo_proceso"] == "M1.1"
    assert registro["mes"] == "Enero"
    assert registro["resultado_obtenido"] == 80.0
    assert registro["resultado_esperado"] == 75.0
    # Las hojas de mejora (Ishikawa, Lewin…) no son hojas de datos: no deben
    # avisar por falta de encabezados. Los avisos de módulos ausentes sí son
    # legítimos aquí, porque el fixture solo tiene M1.
    assert not any("encabezados" in a for a in advertencias)


def test_nombre_de_archivo_es_ascii(session):
    OrganizacionService.actualizar(session, {"nombre": "Municipalidad de Ñuñoa"})
    nombre = ExportService.nombre_archivo(session)
    assert nombre.isascii()
    assert nombre.endswith(".xlsx")
