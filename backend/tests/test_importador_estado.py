import io

import pytest
from openpyxl import Workbook
from sqlmodel import Session, SQLModel, create_engine

from modules.fichas.application.cambio_service import CambioService
from modules.fichas.application.causa_service import CausaService
from modules.fichas.application.ficha_proceso_service import FichaProcesoService
from modules.fichas.application.importador import ImportadorExcel
from modules.fichas.application.importador_estado import ImportadorEstado
from modules.fichas.application.indicador_service import IndicadorService
from modules.fichas.application.inventario_service import InventarioService
from modules.fichas.application.oportunidad_service import OportunidadService
from modules.fichas.application.organizacion_service import OrganizacionService
from modules.plantilla.plantilla_router import _HOJAS_ESTADO, _hoja_datos, _hoja_estado
from modules.procesos.infrastructure.excel_reader import parsear_excel


@pytest.fixture
def session():
    engine = create_engine("sqlite://", connect_args={"check_same_thread": False})
    SQLModel.metadata.create_all(engine)
    with Session(engine) as s:
        yield s


@pytest.fixture
def proceso(session):
    """Un proceso con indicador, como el que dejaría la hoja de datos."""
    InventarioService.crear(session, {"codigo": "M1.1", "nombre": "Identificación"})
    IndicadorService.crear(session, "M1.1", {
        "nombre": "% de requerimientos", "sentido": "Ascendente",
        "unidad": "%", "meta_final": 90,
    })
    return "M1.1"


def _libro(hojas: dict[str, list[list]]):
    """Un xlsx en memoria: {nombre de hoja: [encabezados, fila, …]}."""
    wb = Workbook()
    wb.remove(wb.active)
    for nombre, filas in hojas.items():
        ws = wb.create_sheet(nombre)
        for fila in filas:
            ws.append(fila)
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def _importar(session, hojas):
    org = OrganizacionService.actual(session)
    return ImportadorEstado.importar(session, _libro(hojas), org)


def test_libro_sin_hojas_conocidas_no_hace_nada(session, proceso):
    resumen = _importar(session, {"Datos": [["Codigo", "Mes"], ["M1.1", "Enero"]]})
    assert resumen["omitidas"] == 0
    assert resumen["inventario"] == 0


def test_importa_ficha_sipoc_con_listas_separadas_por_barra(session, proceso):
    _importar(session, {"Fichas SIPOC": [
        ["Codigo", "Tipo", "Dueño", "Objetivo", "Proveedores", "Entradas"],
        ["M1.1", "misional", "Jefe de Planeamiento", "Identificar necesidades",
         "Área usuaria | Mesa de partes", "Solicitud"],
    ]})
    ficha = FichaProcesoService.obtener(session, "M1.1")
    assert ficha["tipo"] == "misional"
    assert ficha["dueno"] == "Jefe de Planeamiento"
    assert ficha["proveedores"] == ["Área usuaria", "Mesa de partes"]
    assert ficha["entradas"] == ["Solicitud"]


def test_importa_causas_y_reconoce_la_raiz(session, proceso):
    _importar(session, {"Ishikawa": [
        ["Codigo", "Categoria (6M)", "Causa", "Es raiz", "Peso"],
        ["M1.1", "Método", "Sin procedimiento estándar", "Sí", 8],
    ]})
    causas = CausaService.listar(session, "M1.1")["ishikawa"]["Método"]
    assert len(causas) == 1
    assert causas[0]["es_raiz"] is True
    assert causas[0]["peso"] == 8


def test_categoria_fuera_de_las_6m_se_omite_sin_abortar(session, proceso):
    resumen = _importar(session, {"Ishikawa": [
        ["Codigo", "Categoria (6M)", "Causa"],
        ["M1.1", "Vibraciones", "Categoría inventada"],
        ["M1.1", "Personas", "Falta capacitación"],
    ]})
    assert resumen["ishikawa"] == 1
    assert resumen["omitidas"] == 1
    assert len(CausaService.listar(session, "M1.1")["ishikawa"]["Personas"]) == 1


def test_fila_de_proceso_inexistente_se_omite(session, proceso):
    resumen = _importar(session, {"Ishikawa": [
        ["Codigo", "Categoria (6M)", "Causa"],
        ["M9.9", "Método", "Proceso que no existe"],
    ]})
    assert resumen["ishikawa"] == 0
    assert resumen["omitidas"] == 1


def test_importa_oportunidad_con_sus_escalas(session, proceso):
    _importar(session, {"Oportunidades": [
        ["Codigo", "Oportunidad", "Accion Propuesta", "Costo (C)", "Impacto (I)",
         "Probabilidad", "Consecuencia", "Estrategia", "Estado"],
        ["M1.1", "Estandarizar el instrumento", "Elaborar guía", 2, 5, 3, 4, "mitigar", "en_curso"],
    ]})
    oportunidades = OportunidadService.listar(session, "M1.1")
    assert len(oportunidades) == 1
    assert oportunidades[0]["factibilidad"] == 10   # C × I
    assert oportunidades[0]["estado"] == "en_curso"


def test_importa_plan_de_cambio_de_lewin(session, proceso):
    _importar(session, {"Gestion del cambio": [
        ["Codigo", "Etapa (Lewin)", "Accion", "Responsable", "Fecha", "Estado"],
        ["M1.1", "Descongelar", "Presentar el diagnóstico", "Jefe", "2026-06-30", "hecho"],
        ["M1.1", "cambiar", "Capacitar al equipo", "Analista", "2026-08-31", "en_curso"],
    ]})
    cambio = CambioService.listar(session, "M1.1")
    assert cambio["progreso"] == {"total": 2, "hechas": 1, "porcentaje": 50.0}


def test_reimportar_actualiza_en_vez_de_duplicar(session, proceso):
    hojas = {"Ishikawa": [
        ["Codigo", "Categoria (6M)", "Causa", "Es raiz", "Peso"],
        ["M1.1", "Método", "Sin procedimiento estándar", "No", 3],
    ]}
    _importar(session, hojas)
    hojas["Ishikawa"][1] = ["M1.1", "Método", "Sin procedimiento estándar", "Sí", 9]
    _importar(session, hojas)

    causas = CausaService.listar(session, "M1.1")["ishikawa"]["Método"]
    assert len(causas) == 1
    assert causas[0]["es_raiz"] is True
    assert causas[0]["peso"] == 9


def test_inventario_completa_producto_y_jerarquia(session, proceso):
    _importar(session, {"Inventario": [
        ["Codigo", "Nivel", "Codigo Padre", "Nombre", "Producto", "Base Legal"],
        ["M1.1", 1, "M1", "Identificación de necesidades", "Informe validado", "D.S. 123"],
    ]})
    proceso_bd = InventarioService._buscar_por_codigo(
        session, OrganizacionService.actual(session).id, "M1.1"
    )
    assert proceso_bd.producto == "Informe validado"
    assert proceso_bd.base_legal == "D.S. 123"
    assert proceso_bd.codigo_padre == "M1"


def test_organizacion_se_toma_del_libro(session, proceso):
    _importar(session, {"Organizacion": [
        ["Nombre", "Sector"],
        ["Municipalidad de Miraflores", "Gobierno local"],
    ]})
    org = OrganizacionService.actual(session)
    assert org.nombre == "Municipalidad de Miraflores"
    assert org.sector == "Gobierno local"


def test_la_plantilla_que_se_descarga_es_importable(session):
    """
    Las filas de ejemplo de la plantilla tienen que entrar tal cual: si alguien
    descarga la plantilla y la sube sin tocarla debe ver datos, no una BD vacía.
    """
    wb = Workbook()
    _hoja_datos(wb)
    for nombre, ayuda, columnas, ejemplos in _HOJAS_ESTADO:
        _hoja_estado(wb, nombre, ayuda, columnas, ejemplos)
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    registros, _ = parsear_excel(buffer)
    ImportadorExcel.importar(session, registros)
    buffer.seek(0)
    resumen = ImportadorEstado.importar(
        session, buffer, OrganizacionService.actual(session)
    )

    assert resumen["omitidas"] == 0
    assert OrganizacionService.actual(session).sector == "Gobierno local"
    assert FichaProcesoService.obtener(session, "M1.1")["tipo"] == "misional"
    assert CambioService.listar(session, "M1.1")["progreso"]["total"] == 3


def test_encabezados_con_titulo_arriba_igual_se_encuentran(session, proceso):
    # Quien edita el libro suele agregar un título sobre la tabla
    _importar(session, {"Ishikawa": [
        ["Diagrama de Ishikawa del proceso M1.1"],
        [],
        ["Codigo", "Categoria (6M)", "Causa"],
        ["M1.1", "Entorno", "Cambios normativos frecuentes"],
    ]})
    assert len(CausaService.listar(session, "M1.1")["ishikawa"]["Entorno"]) == 1
